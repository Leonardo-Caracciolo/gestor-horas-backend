"""
services/export_service.py — Generación de Excel semanal y notificaciones Teams.

El Excel sigue el formato oficial del equipo:
    Columnas: Día | Mes | Año | Nombre | Tipo | ID Proyecto | Descripción | Tarea | Horas

Teams webhook: envía un mensaje card cuando se cierra una semana o se exporta el Excel.
"""
import io
import logging
from datetime import date, datetime, timezone
from decimal import Decimal
from sqlalchemy.orm import Session, joinedload

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from app.models.registro_hora import RegistroHora, EstadoRegistro
from app.models.semana import Semana, EstadoSemana
from app.core.config import get_settings

logger = logging.getLogger(__name__)

# Colores corporativos (paleta del Excel oficial)
_COLOR_HEADER  = "1F4E79"   # Azul oscuro
_COLOR_TOTAL   = "D6E4F0"   # Azul claro
_COLOR_EMPRESA = "2E75B6"   # Azul medio

MESES_ES = {
    1: "Enero", 2: "Febrero", 3: "Marzo", 4: "Abril",
    5: "Mayo", 6: "Junio", 7: "Julio", 8: "Agosto",
    9: "Septiembre", 10: "Octubre", 11: "Noviembre", 12: "Diciembre",
}

DIAS_ES = {
    0: "Lunes", 1: "Martes", 2: "Miércoles", 3: "Jueves",
    4: "Viernes", 5: "Sábado", 6: "Domingo",
}


# ── Estilos ────────────────────────────────────────────────────────────────

def _border() -> Border:
    thin = Side(style="thin")
    return Border(left=thin, right=thin, top=thin, bottom=thin)


def _header_style(cell, bg_color: str = _COLOR_HEADER) -> None:
    cell.font = Font(bold=True, color="FFFFFF", size=10)
    cell.fill = PatternFill("solid", fgColor=bg_color)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = _border()


def _data_style(cell, bold: bool = False) -> None:
    cell.font = Font(bold=bold, size=10)
    cell.alignment = Alignment(vertical="center", wrap_text=True)
    cell.border = _border()


def _total_style(cell) -> None:
    cell.font = Font(bold=True, size=10, color="1F4E79")
    cell.fill = PatternFill("solid", fgColor=_COLOR_TOTAL)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = _border()


# ── Construcción del Excel ─────────────────────────────────────────────────

def generar_excel_semana(
    db: Session,
    semana: Semana,
    solo_aprobados: bool = True,
) -> bytes:
    """
    Genera el Excel semanal con las horas del equipo.

    Formato: una fila por registro, agrupadas por usuario y día.
    Incluye totales por usuario y total general al final.

    Args:
        db: Sesión de BD.
        semana: La semana a exportar.
        solo_aprobados: Si True (default), solo incluye registros Aprobados.
                        Si False, incluye también Enviados (para preview).

    Returns:
        bytes: Contenido del archivo .xlsx.
    """
    estados = [EstadoRegistro.APROBADO]
    if not solo_aprobados:
        estados.append(EstadoRegistro.ENVIADO)

    registros = (
        db.query(RegistroHora)
        .options(
            joinedload(RegistroHora.usuario),
            joinedload(RegistroHora.proyecto),
            joinedload(RegistroHora.ado_task),
        )
        .filter(
            RegistroHora.fecha >= semana.fecha_inicio,
            RegistroHora.fecha <= semana.fecha_fin,
            RegistroHora.estado.in_(estados),
        )
        .order_by(RegistroHora.fecha, RegistroHora.usuario_id, RegistroHora.id)
        .all()
    )

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Semana {semana.fecha_inicio.strftime('%d-%m-%Y')}"

    # ── Cabecera de empresa ────────────────────────────────────────────────
    ws.merge_cells("A1:I1")
    titulo_cell = ws["A1"]
    titulo_cell.value = f"Registro de Horas — Semana del {semana.fecha_inicio.strftime('%d/%m/%Y')} al {semana.fecha_fin.strftime('%d/%m/%Y')}"
    titulo_cell.font = Font(bold=True, color="FFFFFF", size=12)
    titulo_cell.fill = PatternFill("solid", fgColor=_COLOR_EMPRESA)
    titulo_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 22

    # ── Encabezados de columnas ────────────────────────────────────────────
    headers = ["Día", "Mes", "Año", "Nombre", "Tipo", "ID Proyecto", "Descripción", "Tarea", "Horas"]
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=2, column=col, value=header)
        _header_style(cell)

    ws.row_dimensions[2].height = 18

    # ── Datos ──────────────────────────────────────────────────────────────
    row = 3
    total_general = Decimal("0.00")
    usuario_actual = None
    subtotal_usuario = Decimal("0.00")
    subtotal_start_row = 3

    for reg in registros:
        # Insertar subtotal cuando cambia el usuario
        if usuario_actual is not None and reg.usuario_id != usuario_actual:
            _write_subtotal_row(ws, row, usuario_actual_nombre, subtotal_usuario)
            row += 1
            subtotal_usuario = Decimal("0.00")
            subtotal_start_row = row

        usuario_actual = reg.usuario_id
        usuario_actual_nombre = reg.usuario.nombre

        tarea = ""
        if reg.ado_task:
            tarea = f"[{reg.ado_task.ado_id}] {reg.ado_task.titulo[:60]}"
        elif reg.tarea_manual:
            tarea = reg.tarea_manual

        values = [
            DIAS_ES.get(reg.fecha.weekday(), ""),
            MESES_ES.get(reg.fecha.month, ""),
            reg.fecha.year,
            reg.usuario.nombre,
            reg.proyecto.tipo.value if reg.proyecto else "",
            reg.proyecto.id_proyecto_excel if reg.proyecto else "",
            reg.descripcion,
            tarea,
            float(reg.horas),
        ]

        for col, val in enumerate(values, start=1):
            cell = ws.cell(row=row, column=col, value=val)
            _data_style(cell)

        subtotal_usuario += reg.horas
        total_general += reg.horas
        row += 1

    # Último subtotal
    if usuario_actual is not None:
        _write_subtotal_row(ws, row, usuario_actual_nombre, subtotal_usuario)
        row += 1

    # ── Fila de total general ──────────────────────────────────────────────
    ws.merge_cells(f"A{row}:H{row}")
    total_label = ws[f"A{row}"]
    total_label.value = "TOTAL GENERAL"
    _total_style(total_label)

    total_val = ws.cell(row=row, column=9, value=float(total_general))
    _total_style(total_val)

    # ── Anchos de columna ──────────────────────────────────────────────────
    anchos = [10, 12, 6, 24, 12, 16, 50, 40, 8]
    for col, ancho in enumerate(anchos, start=1):
        ws.column_dimensions[get_column_letter(col)].width = ancho

    # ── Freeze panes ──────────────────────────────────────────────────────
    ws.freeze_panes = "A3"

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _write_subtotal_row(ws, row: int, nombre: str, total: Decimal) -> None:
    ws.merge_cells(f"A{row}:H{row}")
    label = ws[f"A{row}"]
    label.value = f"Subtotal — {nombre}"
    _total_style(label)
    val = ws.cell(row=row, column=9, value=float(total))
    _total_style(val)


# ── Cierre de semana ───────────────────────────────────────────────────────

def cerrar_semana(db: Session, semana: Semana) -> Semana:
    """
    Cierra una semana: bloquea nuevas imputaciones y marca como cerrada.

    Raises:
        ValueError: Si la semana ya está cerrada.
    """
    if semana.estado == EstadoSemana.CERRADA:
        raise ValueError(f"La semana del {semana.fecha_inicio} ya está cerrada")
    semana.estado = EstadoSemana.CERRADA
    semana.cerrado_en = datetime.now(timezone.utc)
    db.commit()
    db.refresh(semana)
    return semana


# ── Notificaciones Teams ───────────────────────────────────────────────────

def notificar_teams(mensaje: str, titulo: str = "Gestor de Horas") -> bool:
    """
    Envía una notificación al canal de Teams via webhook.

    El mensaje usa el formato de Adaptive Card de Teams.
    Si TEAMS_WEBHOOK_URL no está configurado, retorna False silenciosamente.

    Args:
        mensaje: Cuerpo del mensaje.
        titulo: Título de la card (default: "Gestor de Horas").

    Returns:
        bool: True si la notificación fue enviada exitosamente.
    """
    import httpx

    settings = get_settings()
    if not settings.TEAMS_WEBHOOK_URL:
        logger.debug("TEAMS_WEBHOOK_URL no configurado, omitiendo notificación")
        return False

    payload = {
        "type": "message",
        "attachments": [
            {
                "contentType": "application/vnd.microsoft.card.adaptive",
                "content": {
                    "$schema": "http://adaptivecards.io/schemas/adaptive-card.json",
                    "type": "AdaptiveCard",
                    "version": "1.4",
                    "body": [
                        {
                            "type": "TextBlock",
                            "text": titulo,
                            "weight": "Bolder",
                            "size": "Medium",
                            "color": "Accent",
                        },
                        {
                            "type": "TextBlock",
                            "text": mensaje,
                            "wrap": True,
                        },
                        {
                            "type": "TextBlock",
                            "text": f"🕒 {datetime.now().strftime('%d/%m/%Y %H:%M')}",
                            "isSubtle": True,
                            "size": "Small",
                        },
                    ],
                },
            }
        ],
    }

    try:
        resp = httpx.post(
            settings.TEAMS_WEBHOOK_URL,
            json=payload,
            timeout=10.0,
        )
        resp.raise_for_status()
        logger.info(f"Notificación Teams enviada: '{titulo}'")
        return True
    except Exception as e:
        logger.warning(f"Error enviando notificación Teams: {e}")
        return False


def notificar_semana_cerrada(db: Session, semana: Semana) -> None:
    """Notifica al equipo que la semana fue cerrada."""
    n_registros = db.query(RegistroHora).filter(
        RegistroHora.fecha >= semana.fecha_inicio,
        RegistroHora.fecha <= semana.fecha_fin,
        RegistroHora.estado == EstadoRegistro.APROBADO,
    ).count()

    notificar_teams(
        titulo="✅ Semana cerrada",
        mensaje=(
            f"La semana del **{semana.fecha_inicio.strftime('%d/%m/%Y')}** "
            f"al **{semana.fecha_fin.strftime('%d/%m/%Y')}** fue cerrada.\n\n"
            f"Registros aprobados: **{n_registros}**"
        ),
    )
